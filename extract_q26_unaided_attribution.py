"""
Extract Q26 Unaided Brand Attribution Data
Table 119: "Which car brand do you believe these elements belong to?" (BEFORE reveal)
"""

from openpyxl import load_workbook
import json

EXCEL_FILE = 'P045556_ALL_Tables_20251020_Private.xlsx'

def extract_q26_attribution():
    """Extract Q26 unaided brand attribution data"""
    wb = load_workbook(EXCEL_FILE, data_only=True)

    # Table 119: Q26 Unaided Brand Attribution
    ws = wb['Table 119']

    attribution_data = []

    # Based on inspection:
    # Row 9: Total
    # Row 10: "Škoda" - correct attribution
    # Row 13: "Other mentions" - competitive confusion
    # Row 16: "Don't know" - no attribution
    # Structure: Every 3 rows (category, percentage, significance)

    categories = {
        10: "Škoda",
        13: "Other mentions",
        16: "Don't know"
    }

    for row_num, brand_category in categories.items():
        # Get counts from the row (Col3-7)
        total_count = ws.cell(row=row_num, column=3).value
        uk_count = ws.cell(row=row_num, column=4).value
        spain_count = ws.cell(row=row_num, column=5).value
        germany_count = ws.cell(row=row_num, column=6).value
        poland_count = ws.cell(row=row_num, column=7).value

        # Get percentages from next row (row_num + 1)
        total_pct = ws.cell(row=row_num + 1, column=3).value
        uk_pct = ws.cell(row=row_num + 1, column=4).value
        spain_pct = ws.cell(row=row_num + 1, column=5).value
        germany_pct = ws.cell(row=row_num + 1, column=6).value
        poland_pct = ws.cell(row=row_num + 1, column=7).value

        # Convert percentages
        def safe_convert_pct(val):
            if val is None or str(val).strip() in ['', '-', '*', '*%']:
                return 0
            if isinstance(val, str):
                val = val.replace('%', '').strip()
                try:
                    return float(val) / 100 if float(val) > 1 else float(val)
                except:
                    return 0
            if isinstance(val, (int, float)):
                return val / 100 if val > 1 else val
            return 0

        attribution_data.append({
            'attribution_category': brand_category,
            'Total_count': int(total_count) if total_count else 0,
            'Total_percent': safe_convert_pct(total_pct),
            'UK_percent': safe_convert_pct(uk_pct),
            'Spain_percent': safe_convert_pct(spain_pct),
            'Germany_percent': safe_convert_pct(germany_pct),
            'Poland_percent': safe_convert_pct(poland_pct)
        })

    return attribution_data

def main():
    print("Extracting Q26 Unaided Brand Attribution data...")

    try:
        attribution_data = extract_q26_attribution()

        # Save to JSON
        output_file = 'q26_unaided_attribution.json'
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(attribution_data, f, indent=2, ensure_ascii=False)

        print(f"✓ Extracted {len(attribution_data)} attribution categories")
        print(f"✓ Saved to {output_file}")

        # Print summary
        print("\nAttribution Categories Found:")
        for item in attribution_data:
            print(f"  - {item['attribution_category']}: {item['Total_percent']:.1%}")

        # Calculate brand clarity score
        skoda_correct = next((item for item in attribution_data if item['attribution_category'] == 'Škoda'), None)
        if skoda_correct:
            print(f"\nBrand Clarity Summary:")
            print(f"  ✓ {skoda_correct['Total_percent']:.0%} correctly identified as Škoda (unaided)")
            print(f"  ✓ Highest clarity: Poland ({skoda_correct['Poland_percent']:.0%}%)")
            print(f"  ✓ Lowest clarity: UK ({skoda_correct['UK_percent']:.0%}%)")

    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
