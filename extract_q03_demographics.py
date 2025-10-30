"""
Extract Q03 associations data with age and gender breakdowns from Excel
"""

from openpyxl import load_workbook
import json

# Load workbook
wb = load_workbook('P045556_ALL_Tables_20251020_Private.xlsx', data_only=True)

# Map Q03 tables to elements (from earlier findings)
q03_tables = {
    18: "Electric Green",
    19: "Facets",
    20: "Type",
    21: "Symbol",
    22: "Sonic",
    23: "Wordmark",
    24: "Emerald Green",
    25: "Hacek",
    26: "Tagline"
}

# Column indices for demographics (confirmed earlier)
MALE_COL = 70
FEMALE_COL = 71
AGE_18_30_COL = 72
AGE_31_42_COL = 73
AGE_43_55_COL = 74

# Initialize data structure
q03_demographics = {}

print("=" * 80)
print("EXTRACTING Q03 ASSOCIATIONS BY AGE/GENDER:")
print("=" * 80)

for table_num, element in q03_tables.items():
    print(f"\nProcessing Table {table_num}: {element}")

    ws = wb[f'Table {table_num}']

    # Get the data rows (starting from row 10, which has first association word)
    # Row 9 has "Total" base size
    base_row = list(ws.iter_rows(min_row=9, max_row=9, values_only=True))[0]

    # Get base sizes for each demographic
    total_base = base_row[1]  # Column B (Total)
    male_base = base_row[MALE_COL]
    female_base = base_row[FEMALE_COL]
    age_18_30_base = base_row[AGE_18_30_COL]
    age_31_42_base = base_row[AGE_31_42_COL]
    age_43_55_base = base_row[AGE_43_55_COL]

    print(f"  Base sizes - Male: {male_base}, Female: {female_base}")
    print(f"  Base sizes - 18-30: {age_18_30_base}, 31-42: {age_31_42_base}, 43-55: {age_43_55_base}")

    # Extract top associations (rows 10-20, first 10 associations)
    associations_by_demo = {
        'gender': {'male': {}, 'female': {}},
        'age': {'18-30': {}, '31-42': {}, '43-55': {}}
    }

    top_words = []

    # Data structure: Row with word/count, followed by row with percentages (which we'll use directly)
    row_num = 10
    while row_num < 100 and len(top_words) < 15:  # Check up to row 100, get top 15
        word_row = list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
        pct_row = list(ws.iter_rows(min_row=row_num+1, max_row=row_num+1, values_only=True))[0]

        word = word_row[1]  # Column B has the association word (not column A!)
        if not word or word == "":
            row_num += 1
            continue

        # Check if next row has percentages (not another word)
        if pct_row[1] is not None and isinstance(pct_row[1], str):
            # This is another word, not a percentage row
            row_num += 1
            continue

        # Use percentages directly from the percentage row
        male_pct = pct_row[MALE_COL] if pct_row[MALE_COL] and isinstance(pct_row[MALE_COL], (int, float)) else 0
        female_pct = pct_row[FEMALE_COL] if pct_row[FEMALE_COL] and isinstance(pct_row[FEMALE_COL], (int, float)) else 0
        age_18_30_pct = pct_row[AGE_18_30_COL] if pct_row[AGE_18_30_COL] and isinstance(pct_row[AGE_18_30_COL], (int, float)) else 0
        age_31_42_pct = pct_row[AGE_31_42_COL] if pct_row[AGE_31_42_COL] and isinstance(pct_row[AGE_31_42_COL], (int, float)) else 0
        age_43_55_pct = pct_row[AGE_43_55_COL] if pct_row[AGE_43_55_COL] and isinstance(pct_row[AGE_43_55_COL], (int, float)) else 0

        # Store percentages
        if male_pct > 0:
            associations_by_demo['gender']['male'][word] = male_pct
        if female_pct > 0:
            associations_by_demo['gender']['female'][word] = female_pct
        if age_18_30_pct > 0:
            associations_by_demo['age']['18-30'][word] = age_18_30_pct
        if age_31_42_pct > 0:
            associations_by_demo['age']['31-42'][word] = age_31_42_pct
        if age_43_55_pct > 0:
            associations_by_demo['age']['43-55'][word] = age_43_55_pct

        top_words.append(word)

        # Move to next word (skip the percentage row)
        row_num += 3  # Skip percentage row + empty row

    print(f"  Extracted {len(top_words)} associations")

    # Store in main data structure
    q03_demographics[element] = associations_by_demo

# Save to JSON
output_file = 'q03_associations_by_demographics.json'
with open(output_file, 'w', encoding='utf-8') as f:
    json.dump(q03_demographics, f, indent=2, ensure_ascii=False)

print("\n" + "=" * 80)
print(f"✓ Saved to {output_file}")
print("=" * 80)

# Print sample
print("\nSample data for Electric Green:")
if "Electric Green" in q03_demographics:
    print("  By Gender (Male):")
    for word, freq in list(q03_demographics["Electric Green"]['gender']['male'].items())[:5]:
        print(f"    {word}: {freq:.1%}")
    print("  By Age (18-30):")
    for word, freq in list(q03_demographics["Electric Green"]['age']['18-30'].items())[:5]:
        print(f"    {word}: {freq:.1%}")
