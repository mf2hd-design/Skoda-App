#!/usr/bin/env python3
"""
Extract Q29 MaxDiff Rankings and Q30 Top-of-Mind Word Associations
from P045556_ALL_Tables_20251020_Private.xlsx

Q29: Tables 123 & 124 - Asset Power Rankings (which elements most strongly linked to Škoda)
Q30: Table 125 - 3 words that come to mind when thinking of Škoda
"""

import json
from openpyxl import load_workbook

# File path
EXCEL_FILE = "P045556_ALL_Tables_20251020_Private.xlsx"

def extract_q29_ranked_first():
    """Extract Q29 Table 123 - Elements ranked 1st (most strongly linked)"""
    print("Extracting Q29 Table 123 (Ranked 1st)...")

    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb['Table 123']

    data = {}

    # Column mapping (based on Excel structure)
    col_map = {
        'C': 'Total',
        'D': 'UK',
        'E': 'Spain',
        'F': 'Germany',
        'G': 'Poland'
    }

    # Start reading from row 9 (after headers)
    # Data is in 3-row pattern: element name/count, percentage, significance
    for row_num in range(9, 38, 3):  # Every 3 rows
        element_cell = ws[f'B{row_num}']
        if not element_cell.value:
            continue

        element_name = str(element_cell.value).strip()

        # Skip if not an element row
        if not element_name or element_name.startswith('Base:'):
            continue

        # Get count row and percentage row
        count_row = row_num
        pct_row = row_num + 1

        element_data = {}

        for col, country in col_map.items():
            # Get count
            count_cell = ws[f'{col}{count_row}']
            count = count_cell.value if count_cell.value else 0

            # Get percentage
            pct_cell = ws[f'{col}{pct_row}']
            pct = pct_cell.value if pct_cell.value else 0

            # Handle string percentages like '*%' or '-'
            if isinstance(pct, str):
                pct = 0

            # Convert percentage to decimal if needed
            if isinstance(pct, (int, float)) and pct > 1:
                pct = pct / 100

            element_data[country] = {
                'count': int(count) if isinstance(count, (int, float)) else 0,
                'percent': float(pct) if isinstance(pct, (int, float)) else 0.0
            }

        data[element_name] = element_data
        print(f"  ✓ {element_name}: {element_data['Total']['percent']:.1%}")

    wb.close()
    return data


def extract_q29_ranked_top3():
    """Extract Q29 Table 124 - Elements ranked in top 3"""
    print("\nExtracting Q29 Table 124 (Ranked Top 3)...")

    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb['Table 124']

    data = {}

    col_map = {
        'C': 'Total',
        'D': 'UK',
        'E': 'Spain',
        'F': 'Germany',
        'G': 'Poland'
    }

    for row_num in range(9, 38, 3):
        element_cell = ws[f'B{row_num}']
        if not element_cell.value:
            continue

        element_name = str(element_cell.value).strip()

        if not element_name or element_name.startswith('Base:'):
            continue

        count_row = row_num
        pct_row = row_num + 1

        element_data = {}

        for col, country in col_map.items():
            count_cell = ws[f'{col}{count_row}']
            count = count_cell.value if count_cell.value else 0

            pct_cell = ws[f'{col}{pct_row}']
            pct = pct_cell.value if pct_cell.value else 0

            # Handle string percentages
            if isinstance(pct, str):
                pct = 0

            if isinstance(pct, (int, float)) and pct > 1:
                pct = pct / 100

            element_data[country] = {
                'count': int(count) if isinstance(count, (int, float)) else 0,
                'percent': float(pct) if isinstance(pct, (int, float)) else 0.0
            }

        data[element_name] = element_data
        print(f"  ✓ {element_name}: {element_data['Total']['percent']:.1%}")

    wb.close()
    return data


def extract_q30_word_associations():
    """Extract Q30 Table 125 - 3 words top of mind"""
    print("\nExtracting Q30 Table 125 (3 Words Top of Mind)...")

    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb['Table 125']

    data = []

    col_map = {
        'C': 'Total',
        'D': 'UK',
        'E': 'Spain',
        'F': 'Germany',
        'G': 'Poland'
    }

    # Start from row 10 (after base count row)
    # Skip the base row (row 9)
    for row_num in range(10, 161, 3):  # Every 3 rows
        word_cell = ws[f'B{row_num}']
        if not word_cell.value:
            continue

        word = str(word_cell.value).strip()

        # Skip empty or metadata rows
        if not word or len(word) < 2:
            continue

        count_row = row_num
        pct_row = row_num + 1

        word_data = {'word': word}

        for col, country in col_map.items():
            count_cell = ws[f'{col}{count_row}']
            count = count_cell.value if count_cell.value else 0

            pct_cell = ws[f'{col}{pct_row}']
            pct = pct_cell.value if pct_cell.value else 0

            # Handle string percentages
            if isinstance(pct, str):
                pct = 0

            if isinstance(pct, (int, float)) and pct > 1:
                pct = pct / 100

            word_data[f'{country}_count'] = int(count) if isinstance(count, (int, float)) else 0
            word_data[f'{country}_percent'] = float(pct) if isinstance(pct, (int, float)) else 0.0

        # Only add if has meaningful data
        if word_data['Total_count'] > 0:
            data.append(word_data)
            print(f"  ✓ {word[:40]}: {word_data['Total_percent']:.1%}")

    wb.close()

    # Sort by total count descending
    data.sort(key=lambda x: x['Total_count'], reverse=True)

    return data


def main():
    """Main extraction function"""
    print("=" * 80)
    print("Q29 & Q30 DATA EXTRACTION")
    print("=" * 80)

    # Extract Q29 data
    q29_first = extract_q29_ranked_first()
    q29_top3 = extract_q29_ranked_top3()

    # Extract Q30 data
    q30_words = extract_q30_word_associations()

    # Save to JSON files
    print("\n" + "=" * 80)
    print("SAVING JSON FILES")
    print("=" * 80)

    with open('q29_rankings_first.json', 'w', encoding='utf-8') as f:
        json.dump(q29_first, f, indent=2, ensure_ascii=False)
    print("✓ Saved: q29_rankings_first.json")

    with open('q29_rankings_top3.json', 'w', encoding='utf-8') as f:
        json.dump(q29_top3, f, indent=2, ensure_ascii=False)
    print("✓ Saved: q29_rankings_top3.json")

    with open('q30_word_associations.json', 'w', encoding='utf-8') as f:
        json.dump(q30_words, f, indent=2, ensure_ascii=False)
    print("✓ Saved: q30_word_associations.json")

    # Print summary
    print("\n" + "=" * 80)
    print("EXTRACTION SUMMARY")
    print("=" * 80)
    print(f"Q29 Ranked 1st: {len(q29_first)} elements")
    print(f"Q29 Top 3: {len(q29_top3)} elements")
    print(f"Q30 Word Associations: {len(q30_words)} unique words/phrases")
    print()
    print("Top 3 Elements (Ranked 1st):")
    sorted_first = sorted(q29_first.items(), key=lambda x: x[1]['Total']['percent'], reverse=True)
    for i, (elem, data) in enumerate(sorted_first[:3], 1):
        print(f"  {i}. {elem}: {data['Total']['percent']:.1%}")

    print("\nTop 5 Word Associations:")
    for i, word_data in enumerate(q30_words[:5], 1):
        print(f"  {i}. {word_data['word']}: {word_data['Total_percent']:.1%} ({word_data['Total_count']} mentions)")

    print("\n✅ Extraction complete!")


if __name__ == '__main__':
    main()
