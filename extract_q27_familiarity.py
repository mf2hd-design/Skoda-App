"""
Extract Q27 Škoda Familiarity Data
Table 120: "Which of the following best describes how familiar you are with Škoda?"
"""

from openpyxl import load_workbook
import json

EXCEL_FILE = 'P045556_ALL_Tables_20251020_Private.xlsx'

def extract_q27_familiarity():
    """Extract Q27 Škoda familiarity scale data"""
    wb = load_workbook(EXCEL_FILE, data_only=True)

    # Table 120: Q27 Škoda Familiarity
    ws = wb['Table 120']

    familiarity_data = []

    # Based on inspection:
    # Row 10: "I know a lot about Škoda"
    # Row 13: "I know some things about Škoda"
    # Row 16: "I have heard the name, but don't know anything about Škoda"
    # Row 19: "I have never heard of Škoda"
    # Row 22: "Don't know"
    # Row 25: "NET: Have heard of Škoda" (aggregate)

    familiarity_levels = {
        10: {"label": "I know a lot about Škoda", "level": "High Knowledge", "order": 1},
        13: {"label": "I know some things about Škoda", "level": "Medium Knowledge", "order": 2},
        16: {"label": "I have heard the name, but don't know anything about Škoda", "level": "Heard Name Only", "order": 3},
        19: {"label": "I have never heard of Škoda", "level": "Never Heard", "order": 4},
        22: {"label": "Don't know", "level": "Unknown", "order": 5}
    }

    for row_num, info in familiarity_levels.items():
        # Get counts from the row (Col3-7)
        total_count = ws.cell(row=row_num, column=3).value
        uk_count = ws.cell(row=row_num, column=4).value
        spain_count = ws.cell(row=row_num, column=5).value
        germany_count = ws.cell(row=row_num, column=6).value
        poland_count = ws.cell(row=row_num, column=7).value

        # Get percentages from next row (row_num + 1)
        total_pct = ws.cell(row=row_num + 1, column=3).value
        uk_pct = ws.cell(row=row_num + 1, column=4).value
        spain_pct = ws.cell(row=row_num + 1, column=5).value
        germany_pct = ws.cell(row=row_num + 1, column=6).value
        poland_pct = ws.cell(row=row_num + 1, column=7).value

        # Convert percentages
        def safe_convert_pct(val):
            if val is None or str(val).strip() in ['', '-', '*', '*%']:
                return 0
            if isinstance(val, str):
                val = val.replace('%', '').strip()
                try:
                    return float(val) / 100 if float(val) > 1 else float(val)
                except:
                    return 0
            if isinstance(val, (int, float)):
                return val / 100 if val > 1 else val
            return 0

        familiarity_data.append({
            'familiarity_label': info['label'],
            'familiarity_level': info['level'],
            'order': info['order'],
            'Total_count': int(total_count) if total_count else 0,
            'Total_percent': safe_convert_pct(total_pct),
            'UK_percent': safe_convert_pct(uk_pct),
            'Spain_percent': safe_convert_pct(spain_pct),
            'Germany_percent': safe_convert_pct(germany_pct),
            'Poland_percent': safe_convert_pct(poland_pct)
        })

    # Add NET aggregate
    net_count = ws.cell(row=25, column=3).value
    net_pct = ws.cell(row=26, column=3).value
    net_uk_pct = ws.cell(row=26, column=4).value
    net_spain_pct = ws.cell(row=26, column=5).value
    net_germany_pct = ws.cell(row=26, column=6).value
    net_poland_pct = ws.cell(row=26, column=7).value

    def safe_convert_pct(val):
        if val is None or str(val).strip() in ['', '-', '*', '*%']:
            return 0
        if isinstance(val, str):
            val = val.replace('%', '').strip()
            try:
                return float(val) / 100 if float(val) > 1 else float(val)
            except:
                return 0
        if isinstance(val, (int, float)):
            return val / 100 if val > 1 else val
        return 0

    familiarity_data.append({
        'familiarity_label': 'NET: Have heard of Škoda',
        'familiarity_level': 'Total Awareness',
        'order': 0,
        'Total_count': int(net_count) if net_count else 0,
        'Total_percent': safe_convert_pct(net_pct),
        'UK_percent': safe_convert_pct(net_uk_pct),
        'Spain_percent': safe_convert_pct(net_spain_pct),
        'Germany_percent': safe_convert_pct(net_germany_pct),
        'Poland_percent': safe_convert_pct(net_poland_pct)
    })

    # Sort by order
    familiarity_data.sort(key=lambda x: x['order'])

    return familiarity_data

def main():
    print("Extracting Q27 Škoda Familiarity data...")

    try:
        familiarity_data = extract_q27_familiarity()

        # Save to JSON
        output_file = 'q27_familiarity.json'
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(familiarity_data, f, indent=2, ensure_ascii=False)

        print(f"✓ Extracted {len(familiarity_data)} familiarity levels")
        print(f"✓ Saved to {output_file}")

        # Print summary
        print("\nFamiliarity Levels Found:")
        for item in familiarity_data:
            print(f"  - [{item['familiarity_level']}] {item['familiarity_label']}: {item['Total_percent']:.1%}")

        # Calculate awareness metrics
        net_awareness = next((item for item in familiarity_data if item['familiarity_level'] == 'Total Awareness'), None)
        high_knowledge = next((item for item in familiarity_data if item['familiarity_level'] == 'High Knowledge'), None)

        if net_awareness and high_knowledge:
            print(f"\nKey Metrics:")
            print(f"  ✓ Total Awareness: {net_awareness['Total_percent']:.0%}")
            print(f"  ✓ High Knowledge: {high_knowledge['Total_percent']:.0%}")
            print(f"  ✓ Highest Awareness Market: Poland ({net_awareness['Poland_percent']:.0%}%)")
            print(f"  ✓ Lowest Awareness Market: Germany ({net_awareness['Germany_percent']:.0%}%)")

    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
