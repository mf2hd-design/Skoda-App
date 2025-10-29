"""
Inspect Table 121 structure to understand Q28 data layout
"""

from openpyxl import load_workbook

EXCEL_FILE = 'P045556_ALL_Tables_20251020_Private.xlsx'

def inspect_table_121():
    """Inspect Table 121 structure"""
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb['Table 121']

    print("=" * 80)
    print("TABLE 121 STRUCTURE INSPECTION")
    print("=" * 80)

    # Print first 50 rows
    for row_num in range(1, 51):
        row_data = []
        for col_num in range(1, 10):  # First 9 columns
            cell_value = ws.cell(row=row_num, column=col_num).value
            if cell_value is not None:
                row_data.append(f"Col{col_num}: {cell_value}")

        if row_data:
            print(f"\nRow {row_num}:")
            print("  " + " | ".join(row_data))

    print("\n" + "=" * 80)

if __name__ == "__main__":
    inspect_table_121()
