"""
Extract QD04 Psychographic/Attitudinal Data
Table 132: "Which of the following statements apply to you?"
"""

from openpyxl import load_workbook
import json

EXCEL_FILE = 'P045556_ALL_Tables_20251020_Private.xlsx'

def extract_qd04_psychographics():
    """Extract QD04 psychographic attitudinal data"""
    wb = load_workbook(EXCEL_FILE, data_only=True)

    # Table 132: QD04 Psychographic Attitudes
    ws = wb['Table 132']

    psychographic_data = []

    # Based on inspection:
    # Row 10: "Lifelong learning is important to me"
    # Row 13: "I enjoy a life filled with challenges, novelty, and change"
    # Row 16: "I have a strong sense of adventure"
    # Row 19: "I value traditional customs and beliefs"
    # Structure: Every 3 rows (statement, percentage, significance)

    statements = {
        10: "Lifelong learning is important to me",
        13: "I enjoy a life filled with challenges, novelty, and change",
        16: "I have a strong sense of adventure",
        19: "I value traditional customs and beliefs"
    }

    for row_num, statement in statements.items():
        # Get counts from the row (Col3-7)
        total_count = ws.cell(row=row_num, column=3).value
        uk_count = ws.cell(row=row_num, column=4).value
        spain_count = ws.cell(row=row_num, column=5).value
        germany_count = ws.cell(row=row_num, column=6).value
        poland_count = ws.cell(row=row_num, column=7).value

        # Get percentages from next row (row_num + 1)
        total_pct = ws.cell(row=row_num + 1, column=3).value
        uk_pct = ws.cell(row=row_num + 1, column=4).value
        spain_pct = ws.cell(row=row_num + 1, column=5).value
        germany_pct = ws.cell(row=row_num + 1, column=6).value
        poland_pct = ws.cell(row=row_num + 1, column=7).value

        # Convert percentages
        def safe_convert_pct(val):
            if val is None or str(val).strip() in ['', '-', '*', '*%']:
                return 0
            if isinstance(val, str):
                val = val.replace('%', '').strip()
                try:
                    return float(val) / 100 if float(val) > 1 else float(val)
                except:
                    return 0
            if isinstance(val, (int, float)):
                return val / 100 if val > 1 else val
            return 0

        # Categorize the statement
        category = categorize_psychographic(statement)

        psychographic_data.append({
            'statement': statement,
            'category': category,
            'Total_count': int(total_count) if total_count else 0,
            'Total_percent': safe_convert_pct(total_pct),
            'UK_percent': safe_convert_pct(uk_pct),
            'Spain_percent': safe_convert_pct(spain_pct),
            'Germany_percent': safe_convert_pct(germany_pct),
            'Poland_percent': safe_convert_pct(poland_pct)
        })

    return psychographic_data

def categorize_psychographic(statement):
    """Categorize psychographic statement into segment type"""
    statement_lower = statement.lower()

    if 'adventure' in statement_lower or 'challenges' in statement_lower or 'novelty' in statement_lower:
        return 'Adventure-Seeking'
    elif 'learning' in statement_lower:
        return 'Learning-Oriented'
    elif 'traditional' in statement_lower:
        return 'Traditional-Values'
    else:
        return 'Other'

def main():
    print("Extracting QD04 Psychographic Attitudinal data...")

    try:
        psychographic_data = extract_qd04_psychographics()

        # Save to JSON
        output_file = 'qd04_psychographics.json'
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(psychographic_data, f, indent=2, ensure_ascii=False)

        print(f"✓ Extracted {len(psychographic_data)} psychographic statements")
        print(f"✓ Saved to {output_file}")

        # Print summary
        print("\nPsychographic Statements Found:")
        for item in psychographic_data:
            print(f"  - [{item['category']}] {item['statement']}: {item['Total_percent']:.1%}")

        # Calculate segment sizes
        print(f"\nSegment Distribution:")
        adventure = sum([item['Total_percent'] for item in psychographic_data if item['category'] == 'Adventure-Seeking']) / 2  # Divide by 2 as we have 2 adventure statements
        learning = next((item['Total_percent'] for item in psychographic_data if item['category'] == 'Learning-Oriented'), 0)
        traditional = next((item['Total_percent'] for item in psychographic_data if item['category'] == 'Traditional-Values'), 0)

        print(f"  ✓ Adventure-Seeking: ~{adventure:.0%}")
        print(f"  ✓ Learning-Oriented: {learning:.0%}")
        print(f"  ✓ Traditional-Values: {traditional:.0%}")

    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
