"""
Extract Q28 Emotional Response to Reveal Data
Table 121-122: "How do you respond knowing these are for Škoda?"
"""

from openpyxl import load_workbook
import json

EXCEL_FILE = 'P045556_ALL_Tables_20251020_Private.xlsx'

def extract_q28_response():
    """Extract Q28 emotional response to brand reveal"""
    wb = load_workbook(EXCEL_FILE, data_only=True)

    # Table 121: Q28 Response to Reveal
    ws = wb['Table 121']

    response_data = []

    # Based on inspection:
    # Row 9: Total
    # Row 10: Response category (Col2=label), Row 11: Percentages (Col3-7)
    # Row 13: Response category, Row 14: Percentages
    # Structure: Every 3 rows (category, percentage, significance)

    for row_num in range(10, 25, 3):  # Rows 10, 13, 16, 19, 22
        # Get response category from Col2
        response_category = ws.cell(row=row_num, column=2).value

        if not response_category or response_category in ['Total', 'Columns Tested']:
            continue

        response_category = str(response_category).strip()

        # Get counts from the same row (Col3-7)
        total_count = ws.cell(row=row_num, column=3).value
        uk_count = ws.cell(row=row_num, column=4).value
        spain_count = ws.cell(row=row_num, column=5).value
        germany_count = ws.cell(row=row_num, column=6).value
        poland_count = ws.cell(row=row_num, column=7).value

        # Get percentages from next row (row_num + 1)
        total_pct = ws.cell(row=row_num + 1, column=3).value
        uk_pct = ws.cell(row=row_num + 1, column=4).value
        spain_pct = ws.cell(row=row_num + 1, column=5).value
        germany_pct = ws.cell(row=row_num + 1, column=6).value
        poland_pct = ws.cell(row=row_num + 1, column=7).value

        # Convert percentages
        def safe_convert_pct(val):
            if val is None or str(val).strip() in ['', '-', '*', '*%']:
                return 0
            if isinstance(val, str):
                val = val.replace('%', '').strip()
                try:
                    return float(val) / 100 if float(val) > 1 else float(val)
                except:
                    return 0
            if isinstance(val, (int, float)):
                return val / 100 if val > 1 else val
            return 0

        total_pct_clean = safe_convert_pct(total_pct)
        uk_pct_clean = safe_convert_pct(uk_pct)
        spain_pct_clean = safe_convert_pct(spain_pct)
        germany_pct_clean = safe_convert_pct(germany_pct)
        poland_pct_clean = safe_convert_pct(poland_pct)

        # Only include if we have valid data
        if total_pct_clean > 0 or total_count:
            response_data.append({
                'response_category': response_category,
                'Total_count': int(total_count) if total_count else 0,
                'Total_percent': total_pct_clean,
                'UK_percent': uk_pct_clean,
                'Spain_percent': spain_pct_clean,
                'Germany_percent': germany_pct_clean,
                'Poland_percent': poland_pct_clean
            })

    return response_data

def categorize_sentiment(response_category):
    """Categorize response into positive/neutral/negative"""
    category_lower = response_category.lower()

    # Positive responses
    if 'fits with what i know and expect' in category_lower:
        return 'Positive'

    # Negative responses
    if 'does not fit' in category_lower or 'not fit' in category_lower:
        return 'Negative'

    # Neutral/Informational
    if 'had not heard' in category_lower or "don't know" in category_lower or 'other' in category_lower:
        return 'Neutral'

    return 'Neutral'

def main():
    print("Extracting Q28 Emotional Response to Reveal data...")

    try:
        response_data = extract_q28_response()

        # Add sentiment categorization
        for item in response_data:
            item['sentiment_category'] = categorize_sentiment(item['response_category'])

        # Save to JSON
        output_file = 'q28_emotional_response.json'
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(response_data, f, indent=2, ensure_ascii=False)

        print(f"✓ Extracted {len(response_data)} response categories")
        print(f"✓ Saved to {output_file}")

        # Print summary
        print("\nResponse Categories Found:")
        for item in response_data:
            print(f"  - {item['response_category']}: {item['Total_percent']:.1%} ({item['sentiment_category']})")

    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
