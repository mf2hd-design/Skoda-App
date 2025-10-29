"""
Inspect Table 119 structure to understand Q26 data layout
Q26: Unaided brand attribution - "Which brand do you think this belongs to?"
"""

from openpyxl import load_workbook

EXCEL_FILE = 'P045556_ALL_Tables_20251020_Private.xlsx'

def inspect_table_119():
    """Inspect Table 119 structure"""
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb['Table 119']

    print("=" * 80)
    print("TABLE 119 STRUCTURE INSPECTION (Q26 - Unaided Brand Attribution)")
    print("=" * 80)

    # Print first 100 rows to understand full structure
    for row_num in range(1, 101):
        row_data = []
        for col_num in range(1, 12):  # First 11 columns
            cell_value = ws.cell(row=row_num, column=col_num).value
            if cell_value is not None:
                row_data.append(f"Col{col_num}: {cell_value}")

        if row_data:
            print(f"\nRow {row_num}:")
            print("  " + " | ".join(row_data))

    print("\n" + "=" * 80)

if __name__ == "__main__":
    inspect_table_119()
