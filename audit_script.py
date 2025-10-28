#!/usr/bin/env python3
"""
Comprehensive Data Verification Audit Script
Compares data in app.py against source Excel files from Savanta
"""

import sys
import json
from pathlib import Path

# Check if openpyxl is available, try to use it
try:
    import openpyxl
    import pandas as pd
    EXCEL_AVAILABLE = True
except ImportError:
    print("WARNING: openpyxl not available. Will extract data structure only.")
    EXCEL_AVAILABLE = False

def extract_excel_structure(excel_path):
    """Extract worksheet names and basic structure without openpyxl"""
    print(f"\n{'='*80}")
    print(f"EXCEL FILE: {excel_path}")
    print(f"{'='*80}")

    if not EXCEL_AVAILABLE:
        print("Cannot read Excel without openpyxl. Skipping detailed extraction.")
        return None

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        print(f"Worksheets found: {wb.sheetnames}")

        data = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n--- Sheet: {sheet_name} ---")
            print(f"Dimensions: {ws.dimensions}")

            # Read first 10 rows to understand structure
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                if i <= 10:
                    rows.append(row)
                if i == 1:
                    print(f"Headers: {row}")

            data[sheet_name] = {
                'dimensions': ws.dimensions,
                'sample_rows': rows
            }

        return data
    except Exception as e:
        print(f"Error reading Excel: {e}")
        return None

def find_q02_q05_tables(excel_path):
    """Find Q02 (recognition) and Q05 (uniqueness) data tables"""
    if not EXCEL_AVAILABLE:
        return None

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        results = {
            'Q02': {},  # Recognition
            'Q05': {}   # Uniqueness
        }

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Search for Q02 or Q05 mentions
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                if row_idx > 100:  # Only check first 100 rows
                    break

                for col_idx, cell_value in enumerate(row):
                    if cell_value and isinstance(cell_value, str):
                        if 'Q02' in cell_value.upper() or 'Q2' in cell_value.upper():
                            print(f"Found potential Q02 reference in {sheet_name} at row {row_idx}: {cell_value}")
                        if 'Q05' in cell_value.upper() or 'Q5' in cell_value.upper():
                            print(f"Found potential Q05 reference in {sheet_name} at row {row_idx}: {cell_value}")

        return results
    except Exception as e:
        print(f"Error searching for Q02/Q05: {e}")
        return None

def main():
    base_path = Path("/Users/ben/Documents/Saffron/Skoda App")

    excel_files = [
        base_path / "P045556_ALL_Tables_20251020_Private.xlsx",
        base_path / "2025-10-06 P045556 - Saffron Brand Assets - Final - V2 - Private.xlsx",
        base_path / "250915_SKO_Ads Overview.xlsx"
    ]

    for excel_file in excel_files:
        if excel_file.exists():
            extract_excel_structure(excel_file)
        else:
            print(f"ERROR: File not found: {excel_file}")

    # Search for Q02 and Q05 tables
    print("\n" + "="*80)
    print("SEARCHING FOR Q02 (RECOGNITION) AND Q05 (UNIQUENESS) DATA")
    print("="*80)
    main_excel = base_path / "P045556_ALL_Tables_20251020_Private.xlsx"
    if main_excel.exists():
        find_q02_q05_tables(main_excel)

if __name__ == "__main__":
    main()
